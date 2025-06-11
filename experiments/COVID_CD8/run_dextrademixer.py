import argparse
import multiprocessing
import os
import warnings
import pickle
from time import strftime
import pandas as pd
import numpyro
import optuna
import numpy as np
from sklearn.metrics import (roc_auc_score, average_precision_score, f1_score, precision_score, recall_score,
                             accuracy_score)

import sys


sys.path.append("../../")
from dextrademixer.model import DextraDemixer
from dextrademixer.utils import convert_str_to_bool_and_none
import muon as mu

multiprocessing.set_start_method("spawn", force=True)


def parse_arguments():
    parser = argparse.ArgumentParser(description="Run tool on multiple simulation files and average results.")
    parser.add_argument("input_files", nargs="+", help="List of input .h5mu files")
    parser.add_argument("output_file", help="Output CSV file for averaged results")
    parser.add_argument("--model_type", default='mixturemodelkmeans', help="Model type",
                        choices=["mixturemodelkmeans", "mixturemodel"])
    parser.add_argument("--mode", type=str, default='I', help="Processing mode")
    parser.add_argument("--pmhc_key", type=str, default="LTDEMIAQY", help="Processing mode")
    parser.add_argument("--gex_key", type=str, default="dex", help="Processing mode")
    parser.add_argument("--neg_ctrl_key", type=str, default=None, help="Negative control parameter")
    parser.add_argument("--ir_clone_key", type=str, default=None, help="Clonotype parameter")
    parser.add_argument("--threads", type=int, default=None, help="Number of threads")
    parser.add_argument("--seed", type=int, default=42, help="Random seed")
    parser.add_argument("--maxiter", type=int, default=5000, help="Upper limit for maxiter for optuna")
    parser.add_argument("--lr", type=float, default=1e-2, help="Learning rate")
    parser.add_argument("--n_trials", type=int, default=5, help="Number of trials, in this case number of seeds")
    parser.add_argument("--mp", type=str, default=True, help="Use multiprocessing")
    parser.add_argument("--alpha_model", type=str, default='kmeans',
                        choices=["overdispersion", "kmeans"],
                        help="Modeling of the alpha parameter. Options: 'overdispersion', 'kmeans'.")
    parser.add_argument("--overdispersion_scale_prior", type=float, default=1,
                        help="Prior for scale parameter of HalfCauchy for overdispersion model. Not used for kmeans.")
    parser.add_argument("--var_hyperprior", type=float, default=1000,
                        help="Prior for scale parameter of kmeans model. Not used for overdispersion.")
    return parser.parse_args()


def run_inference(f_in, args):
    opt_params = {"maxiter": args.maxiter,
                  "nof_inits": 10,
                  "adam": {"init_value": args.lr,},
                  "overdispersion_scale_prior": args.overdispersion_scale_prior,
                  "var_hyperprior": args.var_hyperprior,
                  }

    numpyro.set_host_device_count(1)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        mdata = mu.read(f_in)
    # y_true = mdata.mod["airr"].obs["is_binder"]

    mixer = DextraDemixer(model_type=args.model_type, mode=args.mode, alpha_model=args.alpha_model)
    mixer.preprocess_model_data(mdata,
                                pmhc_key=args.pmhc_key,
                                gex_key=args.gex_key,
                                neg_ctrl_key=args.neg_ctrl_key,
                                ir_clone_key=args.ir_clone_key)

    mixer.model._model_config["overdispersion_scale_prior"] = opt_params["overdispersion_scale_prior"]
    mixer.model._model_config["var_hyperprior"] = opt_params["var_hyperprior"]

    trace, best_loss = mixer.fit_svi(svi_config=opt_params,
                                     nof_inits=opt_params["nof_inits"],
                                     rng_key=args.seed,
                                     return_loss=True)
    p_pred, assignment = mixer.predict_posterior_class(target_fdr=0.05)

    config = (f"{args.model_type}_{args.mode}_{args.neg_ctrl_key}_{args.ir_clone_key}_{args.alpha_model}_{opt_params['overdispersion_scale_prior']},{opt_params['var_hyperprior']}_lr={opt_params['adam']['init_value']}\n"
              f"{f_in.replace('simulation/sim_', '').replace('.h5mu', '')}")

    import pandas as pd
    exp_results = pd.read_excel("../../data/SARS-CoV2/Determined TCR avidity_Katha.xlsx", header=0,
                                sheet_name="Updated_241007_individual_Donor")
    exp_results = exp_results[exp_results.donor != "A29"]
    unreactive_clones = ['10623.0', '4350.0', '2123.0', '1690.0', '9973.0']
    exp_results = exp_results[~exp_results.Clone.isin(unreactive_clones)]
    exp_results = exp_results[exp_results["Clone"].notna()]
    exp_results = exp_results.drop_duplicates("Clone")

    mdata.mod["airr"].obs[f"p_pred"] = p_pred
    mdata.mod["airr"].obs[f"assignment"] = assignment

    from openpyxl import load_workbook

    wb = load_workbook('../../data/SARS-CoV2/Determined TCR avidity_Katha.xlsx')
    ws = wb['Updated_241007_individual_Donor']
    determined_avidity = pd.read_excel("../../data/SARS-CoV2/Determined TCR avidity_Katha.xlsx", header=0,
                                sheet_name="Updated_241007_individual_Donor")
    determined_avidity['label IFNg EC50'] = pd.Series([cell.fill.start_color.index for cell in ws['B:B']][1:]).map(
        {'00000000': 'intermediate avidity', 'FFE7E6E6': 'low avidity', 'FFE2EFD9': 'high avidity'})
    determined_avidity['label mTRBC EC50'] = pd.Series([cell.fill.start_color.index for cell in ws['C:C']][1:]).map(
        {'00000000': 'intermediate avidity', 'FFE7E6E6': 'low avidity', 'FFE2EFD9': 'high avidity'})
    determined_avidity = determined_avidity[determined_avidity['donor'] != "A29"]

    df_airr = mdata.mod["airr"].obs[["clone_id_felix", f"p_pred", f"assignment"]].copy()
    df_airr["LTD"] = mdata.mod["dex"].X[:, 0]
    df_airr["clone_id_felix"] = df_airr["clone_id_felix"].astype("float")
    df_airr = pd.merge(exp_results, df_airr, how="right", left_on="Clone", right_on="clone_id_felix")
    # df_airr = pd.merge(determined_avidity, df_airr, how="right", left_on="Clone", right_on="clone_id_felix")
    df_reactive_clones = pd.merge(exp_results, df_airr, how="inner", left_on="Clone", right_on="clone_id_felix")
    y_true = df_airr["IFNg EC50"].notna()

    clone_precision = len(df_airr[df_airr[f"assignment"] == 1]["clone_id_felix"].unique()) / len(df_airr["clone_id_felix"].unique())
    metrics = {"total": len(df_reactive_clones),
     "TP": np.sum(df_reactive_clones[f"assignment"]),
     "FN": np.sum(df_reactive_clones[f"assignment"] != 1),
     "recall": np.sum(df_reactive_clones[f"assignment"]) / len(df_reactive_clones),
     "clonal_precision": clone_precision}
    print(metrics)

    mixer.plot_results(assignment, p_pred, y_true, args.seed, config)

    np.isnan(df_airr['LTD']).mean()
    np.isnan(mixer.model.data["x"]).mean()


    import seaborn as sns
    import matplotlib.pyplot as plt
    sns.histplot(df_airr, x="LTD", hue=df_airr['IFNg EC50'].notna())
    plt.title('Real binder')
    plt.show()
    os.makedirs("saved_models", exist_ok=True)
    with open(f"saved_models/{config}.pkl", "wb") as f:
        pickle.dump(mixer, f)

    return y_true, p_pred, assignment, best_loss


def worker(f_in, args):
    # If multiprocessing is enabled, the following is needed to track which dataset failed
    if args.mp:
        try:
            y_true, p_pred, assignment_fdr, best_loss = run_inference(f_in, args)
        except Exception as e:
            raise RuntimeError(f"Failed on {f_in}") from e
    else:
        y_true, p_pred, assignment_fdr, best_loss = run_inference(f_in, args)
    return y_true, p_pred, assignment_fdr, best_loss


def main():
    args = parse_arguments()
    args = convert_str_to_bool_and_none(args)

    # Evaluate over multiple datasets
    if args.mp:
        with multiprocessing.Pool(processes=args.threads) as pool:
            results = pool.starmap(worker, [(f_in, args)  # Use trial number as seed
                                            for f_in in args.input_files], chunksize=1)
    else:
        results = []
        for f_in in args.input_files:
            results.append(worker(f_in, args))

    y_true, p_pred, assignment, best_loss = zip(*results)

    results = pd.Series()
    results['roc_auc'] = roc_auc_score(y_true, p_pred)
    results['pr_auc'] = average_precision_score(y_true, p_pred)
    results['f1'] = f1_score(y_true, assignment)
    results['precision'] = precision_score(y_true, assignment)
    results['recall'] = recall_score(y_true, assignment)
    results['accuracy'] = accuracy_score(y_true, assignment)

    results.to_csv(args.output_file)


if __name__ == "__main__":
    main()
